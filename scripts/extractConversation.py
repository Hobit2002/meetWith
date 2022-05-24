import openpyxl,json,os

#loop through tables
for root,dirs,files in os.walk('static/characters/conversations/xlsx'):
    for file in files:
        if "~$" in file:continue
        print("Extracting conversation with %s"%(file)) 
        workbook = openpyxl.load_workbook(os.path.join(root, file))
        sheet = workbook['me']
        #save rows to dict
        dictVersion = {}
        for row in range(2,sheet.max_row+1):
            key = sheet.cell(row = row,column = 1).value
            dictVersion[key] = {}
            dictVersion[key]["text"] =sheet.cell(row = row,column = 2).value
            anacestorStr = "[\"%s\"]"%(sheet.cell(row = row,column = 4).value.replace(",","\",\"")) 
            dictVersion[key]["children"] = json.loads(anacestorStr)
            synonymousCell = sheet.cell(row = row,column = 5).value
            synonyms = key +","+ synonymousCell if synonymousCell != None else key
            dictVersion[key]["synonyms"] = json.loads("[\"%s\"]"%(synonyms.replace(",","\",\"")))
        #save dict to json
        print("Keys:%s"%(dictVersion.keys()))
        jsonFile = open("static/characters/conversations/json/%s.json"%(file[:-5]),"w")
        jsonFile.write(json.dumps(dictVersion))
        jsonFile.close()